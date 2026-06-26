import openpyxl
import io
import sys

# Configure stdout/stderr to use UTF-8 just in case
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

wb = openpyxl.load_workbook("remaxforge_city_pages_seo_v2.xlsx")

with open("excel_contents.md", "w", encoding="utf-8") as f:
    f.write("# Excel File Contents: remaxforge_city_pages_seo_v2.xlsx\n\n")
    
    for sheetname in wb.sheetnames:
        f.write(f"## Sheet: {sheetname}\n\n")
        sheet = wb[sheetname]
        
        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            f.write("Empty sheet.\n\n")
            continue
            
        # Format as markdown table
        # Find maximum columns
        max_cols = max(len(r) for r in rows)
        
        # Write header
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        # Extend headers to max_cols
        if len(headers) < max_cols:
            headers.extend([""] * (max_cols - len(headers)))
        f.write("| " + " | ".join(headers) + " |\n")
        f.write("| " + " | ".join(["---"] * max_cols) + " |\n")
        
        # Write rows
        for row in rows[1:]:
            cells = [str(cell).replace("\n", " <br> ").replace("|", "\\|") if cell is not None else "" for cell in row]
            if len(cells) < max_cols:
                cells.extend([""] * (max_cols - len(cells)))
            f.write("| " + " | ".join(cells) + " |\n")
            
        f.write("\n\n")

print("Successfully written all sheet contents to excel_contents.md")

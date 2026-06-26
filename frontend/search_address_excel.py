import openpyxl
import os

files = ["Remax_Forge_City_Pages_Content_Bible.xlsx", "RemaxForge_SEO_Content_Brief_20Pages.xlsx", "ContentPack.xlsx"]

for file in files:
    if os.path.exists(file):
        try:
            wb = openpyxl.load_workbook(file, read_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    for c_idx, cell in enumerate(row):
                        if cell and ("Kalamboli" in str(cell) or "410218" in str(cell)):
                            print(f"Found in {file} -> Sheet '{sheetname}' -> Row {r_idx}, Col {c_idx}: {cell}")
        except Exception as e:
            print(f"Error reading {file}: {e}")
